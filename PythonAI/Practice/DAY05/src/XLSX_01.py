from openpyxl import Workbook

filename = "../data/sample.xlsx"

# 액셀 파일 생성
wb = Workbook()
ws1 = wb.active
ws1.title="sheet #1"
ws1["A1"] = "Hello"
ws1["A3"] = "안녕하세요"

# 새 시트
ws2 = wb.create_sheet("sheet #2")
ws2["A2"] = "ㅎㅇ"
ws2.cell(3,3).value = "None"
wb.save(filename)
