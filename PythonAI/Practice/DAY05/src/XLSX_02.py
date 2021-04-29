import openpyxl

filename = "../data/sample.xlsx"

# 엑셀 파일 열기
book = openpyxl.load_workbook(filename)

# 시트 추출
sheet = book.worksheets[0]
sheetname = book.sheetnames
print(sheet.title)
print(sheet.max_row,sheet.max_column)