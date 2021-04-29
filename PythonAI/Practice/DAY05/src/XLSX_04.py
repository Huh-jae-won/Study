import openpyxl

filename = "../data/stats_100701.xlsx"

book = openpyxl.load_workbook(filename)
sheet = book.worksheets[0]

max_col = sheet.max_column
max_row = sheet.max_row

# 서울 제외 인구 추출 파일 저장

